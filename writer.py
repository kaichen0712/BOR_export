# writer.py
# ---------------------------------------------------------
# 將 scheduler 產生的排班結果寫入《本月》Sheet
# 結構：
#   - 姓名從 B3 (col=2, row>=3)
#   - 日期從 C1 (col=3, row=1)
#   - 星期在 C2（不依賴）
#
# input: schedule_dict[姓名][YYYY-MM-DD] = 班別字串
# ---------------------------------------------------------

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, NamedStyle, Alignment


BASE_DIR = Path(__file__).resolve().parent


def _load_dates_from_json(filename):
    path = BASE_DIR / filename
    try:
        with open(path, "r", encoding="utf-8") as f:
            return set(json.load(f))
    except FileNotFoundError:
        return set()


def _to_halfwidth(text: str) -> str:
    """將全形英數轉半形，保留其他字元"""
    if not text:
        return text
    return text.translate(
        str.maketrans(
            {chr(full): chr(full - 0xFEE0) for full in range(ord("！"), ord("～") + 1)}
        )
    )


def _clean_name(name: str) -> str:
    """移除姓名前後的英文、數字、空白（含全形）、星號；移除所有空白"""
    if not name:
        return ""
    name = _to_halfwidth(str(name))
    name = name.replace("\u3000", " ")  # 全形空白轉半形
    name = re.sub(r'^[A-Za-z0-9\s＊\*]+', "", name)
    name = re.sub(r'[A-Za-z0-9\s＊\*]+$', "", name)
    name = re.sub(r'[\s\u3000]+', "", name)  # 移除所有空白
    return name.strip()


def write_schedule_to_excel(schedule_dict, template_path, output_path, base_year, target_month, monthly_rules_applied=None):
    """
    寫入排班到 Excel
    
    參數：
    - monthly_rules_applied: {name: [rules...]} 記錄套用月份規則（P1/P2/換心）的人員
      這些人員的班別會用黑字填入（預設規則）
    """
    if monthly_rules_applied is None:
        monthly_rules_applied = {}
    
    # -----------------------------------------------------
    # 1) 開啟模板
    # -----------------------------------------------------
    wb = openpyxl.load_workbook(template_path)
    ws = wb["test"]

    # -----------------------------------------------------
    # 1.5) 定義字體樣式
    # -----------------------------------------------------
    # 標楷體、字體大小 12（一般文字）
    base_font = Font(name="標楷體", size=12)
    # 標楷體、字體大小 12、紅色（寫入的班別資料）
    red_font = Font(name="標楷體", size=12, color="FF0000")
    # 假日字體（紅色）
    holiday_font = Font(name="標楷體", size=12, color="FF0000")
    
    # 定義對齊方式（水平垂直置中）
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # -----------------------------------------------------
    # 1.6) 設定整個工作表的字體為標楷體 12，並水平垂直置中
    # -----------------------------------------------------
    for row in ws.iter_rows():
        for cell in row:
            cell.font = base_font
            cell.alignment = center_alignment

    # -----------------------------------------------------
    # 1.7) 記錄原本就有值的儲存格位置（這些要保留黑色）
    #      同時處理原本值中的空白字元換行（日期列除外）
    #      如果原本的值是休假類，改成紅色
    # -----------------------------------------------------
    leave_types = ["休假", "特別休假", "例假", "休息日", "國定假"]
    original_cells = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                original_cells.add((cell.row, cell.column))
                
                cell_str = str(cell.value).strip()
                
                # 如果是休假類，改成紅色
                if cell_str in leave_types:
                    cell.font = red_font
                
                # 如果原本的值包含空白字元，替換為換行並設定自動換行
                # 但日期列（row=1）不處理
                if cell.row > 1:
                    if re.search(r'\s', cell_str):
                        cell.value = re.sub(r'\s+', '\n', cell_str)
                        cell.alignment = wrap_center_alignment
                    
    print(f"[DEBUG] 原本有值的儲存格數量：{len(original_cells)}")

    # -----------------------------------------------------
    # 2) 解析日期欄位（C1 往右）
    # -----------------------------------------------------
    date_row = 1  # 日期列在第 1 列
    date_col_start = 3  # Column C = 3
    weekend_dates = _load_dates_from_json("weekend.json")
    holiday_dates = _load_dates_from_json("holidays.json")

    # 1.5) 先把 schedule_dict 的姓名清洗並合併（避免同人多種寫法）
    normalized_schedule = {}
    for raw_name, date_map in schedule_dict.items():
        name = _clean_name(raw_name)
        if not name:
            continue
        if name not in normalized_schedule:
            normalized_schedule[name] = {}
        for k, v in date_map.items():
            normalized_schedule[name][k] = v
    schedule_dict = normalized_schedule

    excel_dates = {}  # { "YYYY-MM-DD": col_index }
    col = date_col_start

    while True:
        cell = ws.cell(row=date_row, column=col).value
        if cell is None:
            break

        if isinstance(cell, datetime):
            date_key = cell.strftime("%Y-%m-%d")
            excel_dates[date_key] = col
        else:
            try:
                dt = pd.to_datetime(cell)
                date_key = dt.strftime("%Y-%m-%d")
                excel_dates[date_key] = col
            except Exception:
                pass

        col += 1
    print(f"[DEBUG] 解析到日期欄位 {len(excel_dates)} 個 → {sorted(list(excel_dates.keys()))}")

    # -----------------------------------------------------
    # 3) 解析姓名列（從 row 3 開始，col=2）
    # -----------------------------------------------------
    name_col = 2  # B column
    name_row_start = 3
    row = name_row_start

    name_to_row = {}  # {姓名: row_index}

    while True:
        cell = ws.cell(row=row, column=name_col).value
        if cell is None:
            break
        name = _clean_name(cell)
        if name:
            name_to_row[name] = row
        row += 1

    last_name_row = row - 1
    print(f"[DEBUG] 解析到姓名 {len(name_to_row)} 個 → {list(name_to_row.keys())}")

    # 3.5) 先把例假/國定假寫滿整欄（稍後班別會覆蓋）
    #      原本有值的儲存格完全不覆蓋
    for date_key, col_index in excel_dates.items():
        if date_key in holiday_dates:
            label = "國定假"
        elif date_key in weekend_dates:
            label = "例假"
        else:
            continue

        for r_idx in range(name_row_start, last_name_row + 1):
            # 原本有值的儲存格完全跳過，不覆蓋
            if (r_idx, col_index) in original_cells:
                continue
            cell = ws.cell(row=r_idx, column=col_index)
            cell.value = label
            cell.font = holiday_font
            # 國定假較長，需要自動換行
            if label == "國定假":
                cell.alignment = wrap_center_alignment
            else:
                cell.alignment = center_alignment
    print("[DEBUG] 已預填例假/國定假完畢（跳過原本有值的儲存格）")

    # -----------------------------------------------------
    # 4) 將班別寫入對應位置
    #    原本有值的儲存格完全不覆蓋，以 template 資料為準
    #    套用月份規則（P1/P2/換心）的人員使用黑字
    #    其他新寫入的班別使用紅字
    #    「特別休假」等長文字啟用自動換行
    # -----------------------------------------------------
    write_count = 0
    skip_count = 0
    rule_count = 0  # 月份規則寫入計數
    
    for raw_name, date_map in schedule_dict.items():
        name = _clean_name(raw_name)
        if name not in name_to_row:
            print(f"[WARN] 找不到姓名：{raw_name}（清洗後：{name}），略過")
            continue

        user_row = name_to_row[name]
        
        # 判斷該員工是否套用了月份規則（P1/P2/換心）
        is_rule_person = name in monthly_rules_applied

        for date_key, shift in date_map.items():
            if date_key not in excel_dates:
                print(f"[WARN] 找不到日期欄位：{date_key}（姓名: {name}），略過")
                continue
            user_col = excel_dates[date_key]

            # 原本有值的儲存格完全跳過，不覆蓋
            if (user_row, user_col) in original_cells:
                skip_count += 1
                continue

            cell = ws.cell(row=user_row, column=user_col)
            
            # 處理班別值：如果包含任何空白字元，替換為換行符號
            cell_value = shift
            if re.search(r'\s', shift):
                cell_value = re.sub(r'\s+', '\n', shift)
            cell.value = cell_value
            
            # 決定字體顏色：
            # - 月份規則人員的「工作班別」（7~3 等）用黑字
            # - 休假類（休假、特別休假、例假、休息日、國定假）都用紅字
            work_shifts = ["7~3", "3~11", "23~7"]
            is_work_shift = shift in work_shifts
            
            if is_rule_person and is_work_shift:
                cell.font = base_font  # 黑字（月份規則的工作班別）
                rule_count += 1
            else:
                cell.font = red_font  # 紅字（休假類或非月份規則人員）
            
            # 對於「特別休假」、「國定假」、「休息日」或包含空白的班別啟用自動換行
            needs_wrap = shift in ["特別休假", "國定假", "休息日"] or re.search(r'\s', shift)
            if needs_wrap:
                cell.alignment = wrap_center_alignment
            else:
                cell.alignment = center_alignment
            
            write_count += 1

    print(f"[DEBUG] 寫入班別筆數：{write_count}（月份規則黑字：{rule_count}），跳過（原本有值）：{skip_count}")

    # -----------------------------------------------------
    # 5) 輸出檔案
    # -----------------------------------------------------
    wb.save(output_path)
    print(f"✔ 已成功寫入排班 → {output_path}")


def write_schedule_to_excel_memory(schedule_dict, base_year, target_month, monthly_rules_applied=None, identity_map=None):
    """
    直接在記憶體中產生排班 Excel（不需要模板）
    
    返回：BytesIO 物件，可直接用於下載
    """
    import io
    import calendar
    
    if monthly_rules_applied is None:
        monthly_rules_applied = {}
    if identity_map is None:
        identity_map = {}
    
    # 建立新的 workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排班表"
    
    # 定義字體樣式
    base_font = Font(name="標楷體", size=12)
    red_font = Font(name="標楷體", size=12, color="FF0000")
    header_font = Font(name="標楷體", size=12, bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 載入假日資料
    weekend_dates = _load_dates_from_json("weekend.json")
    holiday_dates = _load_dates_from_json("holidays.json")
    
    # 計算該月天數
    _, days_in_month = calendar.monthrange(base_year, target_month)
    
    # 星期對照
    weekday_names = ['一', '二', '三', '四', '五', '六', '日']
    
    # =============================================
    # 1. 建立表頭
    # =============================================
    # A1: 空白, B1: "姓名"
    ws.cell(row=1, column=1).value = ""
    ws.cell(row=1, column=2).value = "姓名"
    ws.cell(row=1, column=2).font = header_font
    ws.cell(row=1, column=2).alignment = center_alignment
    
    # A2: 空白, B2: "星期"
    ws.cell(row=2, column=1).value = ""
    ws.cell(row=2, column=2).value = "星期"
    ws.cell(row=2, column=2).font = header_font
    ws.cell(row=2, column=2).alignment = center_alignment
    
    # C1 往右：日期
    # C2 往右：星期
    date_to_col = {}  # {YYYY-MM-DD: column}
    for day in range(1, days_in_month + 1):
        col = 2 + day  # C=3, D=4, ...
        date_obj = datetime(base_year, target_month, day)
        date_key = date_obj.strftime("%Y-%m-%d")
        
        # 日期列
        ws.cell(row=1, column=col).value = date_obj
        ws.cell(row=1, column=col).number_format = 'M/D'
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = center_alignment
        
        # 星期列
        wd = date_obj.weekday()
        ws.cell(row=2, column=col).value = weekday_names[wd]
        ws.cell(row=2, column=col).font = header_font
        ws.cell(row=2, column=col).alignment = center_alignment
        
        # 週末或假日標紅
        if date_key in weekend_dates or date_key in holiday_dates:
            ws.cell(row=1, column=col).font = Font(name="標楷體", size=12, bold=True, color="FF0000")
            ws.cell(row=2, column=col).font = Font(name="標楷體", size=12, bold=True, color="FF0000")
        
        date_to_col[date_key] = col
    
    # =============================================
    # 2. 填入人員姓名和排班
    # =============================================
    leave_types = ["休假", "特別休假", "例假", "休息日", "國定假"]
    work_shifts = ["7~3", "3~11", "23~7"]
    
    row = 3
    for name, date_map in schedule_dict.items():
        # 如果是用戶指定的順序，就不過濾；否則過濾非身分表人員
        # （這裡不再過濾，因為 app.py 已經處理好順序了）
        # B 欄：姓名
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=2).font = base_font
        ws.cell(row=row, column=2).alignment = center_alignment
        
        # 身分別
        identity = identity_map.get(name, '')  # 找不到就留空
        
        # A 欄：身分（可選）
        ws.cell(row=row, column=1).value = identity
        ws.cell(row=row, column=1).font = base_font
        ws.cell(row=row, column=1).alignment = center_alignment
        
        # 是否套用月份規則
        is_rule_person = name in monthly_rules_applied
        
        # 填入每天的班別
        for day in range(1, days_in_month + 1):
            date_obj = datetime(base_year, target_month, day)
            date_key = date_obj.strftime("%Y-%m-%d")
            col = date_to_col.get(date_key)
            
            if col is None:
                continue
            
            cell = ws.cell(row=row, column=col)
            shift = date_map.get(date_key, "")
            
            if shift:
                cell.value = shift
                
                # 決定字體顏色
                is_work_shift = shift in work_shifts
                
                if is_rule_person and is_work_shift:
                    cell.font = base_font  # 黑字（月份規則的工作班別）
                elif shift in leave_types:
                    cell.font = red_font  # 紅字（休假類）
                else:
                    cell.font = red_font  # 紅字（其他新寫入）
                
                # 自動換行
                if shift in ["特別休假", "國定假", "休息日"]:
                    cell.alignment = wrap_center_alignment
                else:
                    cell.alignment = center_alignment
            else:
                cell.alignment = center_alignment
        
        row += 1
    
    # =============================================
    # 3. 設定欄寬
    # =============================================
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    for day in range(1, days_in_month + 1):
        col_letter = openpyxl.utils.get_column_letter(2 + day)
        ws.column_dimensions[col_letter].width = 7
    
    # =============================================
    # 4. 輸出到記憶體
    # =============================================
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    print(f"✔ 已成功產生排班表（{base_year}年{target_month}月）")
    return output


# 單獨測試用（可刪）
if __name__ == "__main__":
    dummy = {
        "江載仁": {
            "2026-02-04": "7~3",
            "2026-02-05": "3~11",
        }
    }
    write_schedule_to_excel(dummy, "11502.xlsx", "output_11502.xlsx", 2026, 2)